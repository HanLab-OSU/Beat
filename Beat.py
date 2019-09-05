from __future__ import division
from Bio import SeqIO
from Bio.Seq import Seq
import matplotlib.pyplot as plt
from matplotlib.table import table
import matplotlib.colors as mcolors
import sys
import os
import pandas as pd
import timeit
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from matplotlib.font_manager import FontProperties
import numpy as np
from scipy import stats

start = timeit.default_timer()
# spacer_err = []    # track the file which contain no spacer
print("running...")

#############################################################
# Beat8 - Base Editing Analysis Tool version 8
# Version 8 finds the noise and filter the outliers from the noise
# Run batch analysis as python Beat8.py ./example/template1.csv
# or python Beat8.py foldername all GGgtaaatgtagttgataat 6 AG
# to analyze all sequencing files if they have the same spacer
# or Analyze individual sequencing file as
# python Beat8.py foldername sample.ab1 TCGGCCACCACAGGGAAGCT 6 CT
# The editing effiency is determined after substracting
# the background noise without normalization to controls
#############################################################

# define degenerate nucleotides


A = {"A"}
C = {"C"}
G = {"G"}
T = {"T"}

letter2nucl = {
    "A": A,
    "C": C,
    "G": G,
    "T": T,
    "R": A | G,
    "Y": C | T,
    "S": G | C,
    "W": A | T,
    "K": G | T,
    "M": A | C,
    "B": C | G | T,
    "D": A | G | T,
    "H": A | C | T,
    "V": A | C | G,
    "N": {"A", "C", "G", "T"}
    }


def rev_comp(seq):
    return str(Seq(seq).reverse_complement())

def removeOutliersMAD(x):
    # Median absolute deviation (MAD), based on the median, is a robust non-parametric statistics.
    # https://en.wikipedia.org/wiki/Median_absolute_deviation.
    x_median = np.median(x)
    resultList = np.copy(x)
    mad = 1.4826 * np.median(np.abs(x - x_median))
    resultList[(np.abs(x - x_median) > 3 * mad)] = x_median
    return resultList.tolist()


def removeOutliersZ(x):
    # The Z-score is the signed number of standard deviations by which the value of an observation
    # or data point is above the mean value of what is being observed or measured.
    z = np.abs(stats.zscore(x))
    resultList = []

    # See if we can del the for loop here
    for y in np.where(z ** 2 < 3):
        for yy in y:
            resultList.append(x[yy])

    return resultList


def removeOutliers(x, outlierConstant):
    # The interquartile range (IQR), also called the midspread or middle 50%, or technically H-spread,
    # is a measure of statistical dispersion, being equal to the difference between 75th and 25th
    # percentiles, or between upper and lower quartiles, IQR = Q3 - Q1.
    # It is a measure of the dispersion similar to standard deviation or variance,
    # but is much more robust against outliers.
    a = np.array(x)
    upper_quartile = np.percentile(a, 75)
    lower_quartile = np.percentile(a, 25)
    IQR = (upper_quartile - lower_quartile) * outlierConstant
    resultList = []
    for y in a.tolist():
        if lower_quartile - IQR >= y <= upper_quartile + IQR:
            resultList.append(y)
    return resultList


def make_colormap(seq):
    """Return a LinearSegmentedColormap
    seq: a sequence of floats and RGB-tuples. The floats should be increasing
    and in the interval (0,1).
    """
    seq = [(None,) * 3, 0.0] + list(seq) + [1.0, (None,) * 3]
    cdict = {'red': [], 'green': [], 'blue': []}
    for i, item in enumerate(seq):
        if isinstance(item, float):
            r1, g1, b1 = seq[i - 1]
            r2, g2, b2 = seq[i + 1]
            cdict['red'].append([item, r1, r2])
            cdict['green'].append([item, g1, g2])
            cdict['blue'].append([item, b1, b2])
    return mcolors.LinearSegmentedColormap('CustomMap', cdict)


def subseq(seq1, seq2):
    l1 = len(seq1)
    l2 = len(seq2)
    pattern = [letter2nucl[letter] for letter in seq1]
    basecall = [letter2nucl[letter] for letter in seq2]
    i = 0
    spacerpos = -1
    while i < l2 - l1 + 1:
        sub_flag = True
        for j, nucl in enumerate(pattern):
            if not (basecall[i + j] & nucl):
                # empty set intersection
                sub_flag = False
                break
        if sub_flag:  # find all subsets or only first/last one.
            spacerpos = i
        i += 1

    return spacerpos


def find_subseq(pattern, seq, base_change, base_pos_in_spacer):
    # Add a base_change check
    strand = ''
    if base_change == 'AG':
        pos_in_call = subseq(pattern.replace('A','R'),seq)
    if base_change == 'CT':
        pos_in_call = subseq(pattern.replace('C', 'Y'), seq)
    if pos_in_call == -1:
        pattern = rev_comp(pattern)
        base_pos_in_spacer = len(pattern) - base_pos_in_spacer + 1
        base_change = rev_comp(base_change[0]) + rev_comp(base_change[1])
        if base_change == 'TC':
            pos_in_call = subseq(pattern.replace('T', 'Y'), seq)
        if base_change == 'GA':
            pos_in_call = subseq(pattern.replace('G', 'R'), seq)
        if pos_in_call != -1:
            strand = '-'
    else:
        strand = '+'

    return pos_in_call, strand, base_pos_in_spacer


def background_estimation(base_call, peaks_vals, trace):

    Gbg_vals = []
    Abg_vals = []
    Tbg_vals = []
    Cbg_vals = []
    for i in range(100, len(base_call) - 50):
        x = base_call[i]
        base_pos = peaks_vals[i]
        if x in letter2nucl["N"]:
            if x != "A":
                Abg_vals.append(max(trace['A'][base_pos - 3: base_pos + 3]))
            if x != "G":
                Gbg_vals.append(max(trace['G'][base_pos - 3: base_pos + 3]))
            if x != "C":
                Cbg_vals.append(max(trace['C'][base_pos - 3: base_pos + 3]))
            if x != "T":
                Tbg_vals.append(max(trace['T'][base_pos - 3: base_pos + 3]))

    Gbg_vals_cleaned = removeOutliersMAD(Gbg_vals)
    Abg_vals_cleaned = removeOutliersMAD(Abg_vals)
    Tbg_vals_cleaned = removeOutliersMAD(Tbg_vals)
    Cbg_vals_cleaned = removeOutliersMAD(Cbg_vals)
    #print([Gbg_vals,Gbg_vals,Tbg_vals,Cbg_vals])
    #print([Gbg_vals_cleaned,Abg_vals_cleaned,Tbg_vals_cleaned,Cbg_vals_cleaned])
    bg_dict = {
        'G': round(sum(Gbg_vals_cleaned) / len(Gbg_vals_cleaned),1),
        'A': round(sum(Abg_vals_cleaned) / len(Abg_vals_cleaned),1),
        'T': round(sum(Tbg_vals_cleaned) / len(Tbg_vals_cleaned),1),
        'C': round(sum(Cbg_vals_cleaned) / len(Cbg_vals_cleaned),1)
    }
    return bg_dict


def write_excel(data, spacer, filename, spacer_pos, mutated_base, estimation, strand, peak_intensity):

    # Excel output Construction
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Convert to 1 digit
    for i in data[1:]:
        for j in i:
            round(j,1)
    # Rows can also be appended
    for row in data:
        ws.append(row)
    maxRow = ws.max_row
    maxCol = ws.max_column

    # Format header row
    for cell in ws["1:1"]:
        cell.font = Font(size=12, color='00000000', italic=False, bold=True)
        cell.fill = PatternFill(fgColor='D3D3D3', fill_type='solid')

    # Format data row
    for x in range(2, maxRow + 1):
        for y in range(1, maxCol):
            ws.cell(row=x, column=y).font = Font(size=12, bold=False)

    # Highlight the mutated base
    ws.cell(row=1, column=mutated_base).font = Font(size=12, color='FF0000', bold=True)

    # Highlight the entire column of the mutated base
    for x in range(1, maxRow + 1):
        ws.cell(row=x, column=mutated_base).fill = PatternFill(fgColor='FFEE08', fill_type='solid')

    # Insert a column in the beginning
    ws.insert_cols(1)
    '''simplify this part'''
    # Add base index for each row
    ws.cell(row=1, column=1).value = 'Spacer'
    ws.cell(row=1, column=1).font = Font(size=12, color='000000', bold=True)
    ws.cell(row=2, column=1).value = 'G'
    ws.cell(row=2, column=1).font = Font(size=12, color='000000', bold=True)
    ws.cell(row=3, column=1).value = 'A'
    ws.cell(row=3, column=1).font = Font(size=12, color='008000', bold=True)
    ws.cell(row=4, column=1).value = 'T'
    ws.cell(row=4, column=1).font = Font(size=12, color='FF0000', bold=True)
    ws.cell(row=5, column=1).value = 'C'
    ws.cell(row=5, column=1).font = Font(size=12, color='0000FF', bold=True)
    for x in range(1, maxRow + 1):
        ws.cell(row=x, column=1).fill = PatternFill(fgColor='D3D3D3', fill_type='solid')

    # Insert information after the table
    ws.cell(row=7, column=1).value = filename
    if strand == "+":
        ws.cell(row=8, column=1).value = 'spacer: ' + spacer
    else:
        ws.cell(row=8, column=1).value = 'spacer: ' + rev_comp(spacer)
    ws.cell(row=9, column=1).value = 'spacer position in the base calls: ' + str(spacer_pos + 1)
    ws.cell(row=10, column=1).value = 'Estimated background: G:'+str(estimation['G'])+'; A:'+str(estimation['A'])+\
                                      '; T:'+str(estimation['T'])+'; C:'+str(estimation['C'])
    ws.cell(row=11, column=1).value = 'Peak values at position ' + str(mutated_base) + ', G:' + str(
        peak_intensity['G'][mutated_base - 1]) + '; A:' + str(peak_intensity['A'][mutated_base - 1]) + '; T: ' + str(
        peak_intensity['T'][mutated_base - 1]) + '; C: ' + str(peak_intensity['C'][mutated_base - 1])

    # Save the file
    wb.save('./result/'+filename.split('.')[0]+'.xlsx')


def plt_figure(trace, peaks_vals, spacer, pos_in_call, filename, pos_list, base_pos_in_spacer,strand):
    # Plot the trace and table showing % of each base along the spacer

    fig, (ax1, ax2) = plt.subplots(2, figsize=(4, 2.5))
    fig.patch.set_visible(False)
    ax1 = plt.subplot2grid((2, 1), (0, 0))
    ax2 = plt.subplot2grid((2, 1), (1, 0))
    Arialfont = {'fontname': 'Arial'}
    ax1.plot(trace['G'][peaks_vals[pos_in_call] - 8: peaks_vals[pos_in_call + len(spacer) - 1] + 8], color='black',
             label='G', linewidth=0.6)
    ax1.plot(trace['A'][peaks_vals[pos_in_call] - 8: peaks_vals[pos_in_call + len(spacer) - 1] + 8], color='g', label='A',
             linewidth=0.6)  # A
    ax1.plot(trace['T'][peaks_vals[pos_in_call] - 8: peaks_vals[pos_in_call + len(spacer) - 1] + 8], color='r', label='T',
             linewidth=0.6)  # T
    ax1.plot(trace['C'][peaks_vals[pos_in_call] - 8: peaks_vals[pos_in_call + len(spacer) - 1] + 8], color='b', label='C',
             linewidth=0.6)  # C
    ax1.annotate('G', xy=(10, 60), xycoords='axes points', size=5, color='black', weight='bold', ha='right', va='top')
    ax1.annotate('A', xy=(14, 60), xycoords='axes points', size=5, color='g', weight='bold', ha='right', va='top')
    ax1.annotate('T', xy=(18, 60), xycoords='axes points',
                 size=5, color='r', weight='bold', ha='right', va='top')
    ax1.annotate('C', xy=(22, 60), xycoords='axes points',
                 size=5, color='b', weight='bold', ha='right', va='top')
    ax1.axis('off')
    ax1.axis('tight')
    ax1.set_title(filename.split(".")[0], fontsize=8, **Arialfont)

    # Put a legend to the right of the current axis
    # plt.legend([a_plot, c_plot, g_plot, t_plot], ['A', 'C', 'G', 'T'], bbox_to_anchor=(0.98,0.5), loc="center left", prop={'size': 6}, borderaxespad=0., ncol=1)

    ax2.axis('off')
    ax2.axis('tight')
    df = pd.DataFrame(pos_list, columns=list(spacer))
    xx = (df.values * 2.56).astype(int)
    c = mcolors.ColorConverter().to_rgb
    mycm = make_colormap(
        [c('white'), c('yellow'), 0.3, c('yellow'), c('red'), 0.60, c('red'), c('green'), 0.8, c('green')])

    # colours=plt.cm.CMRmap(xx)
    colours = plt.cm.get_cmap(mycm)(xx)
    stats_table = table(ax2, cellText=df.values, rowLabels='GATC', colLabels=df.columns, loc='center',
                        cellColours=colours)
    stats_table._cells[(0, base_pos_in_spacer - 1)]._text.set_color('red')
    stats_table._cells[(2, -1)]._text.set_color('green')
    stats_table._cells[(3, -1)]._text.set_color('red')
    stats_table._cells[(4, -1)]._text.set_color('blue')
    # stats_table._cells[(1,base_pos_in_spacer-1)].set_facecolor('red')

    for (row, col), cell in stats_table.get_celld().items():
        cell.set_linewidth(0)
        if (col == -1):
            cell.set_text_props(fontproperties=FontProperties(family='Arial', size=8, weight='bold'))
        elif (row == 0):
            cell.set_text_props(fontproperties=FontProperties(family='Arial', size=8))
        else:
            cell.set_text_props(fontproperties=FontProperties(family='Arial', size=6))

            # shrink current axis
    box = ax2.get_position()
    ax2.set_position([box.x0 + 0.038, box.y0 + 0.15, box.width * 0.9, box.height * 0.8])
    # ax2.arrow(0, 20, 10, 47, head_width=1, head_length=2, fc='k', ec='k')
    if (strand == "+"):
        ax2.annotate("5'>", xy=(0, 47), xycoords='axes points',
                     size=5, color='k', weight='bold', ha='right', va='top')
    else:
        ax2.annotate("3'<", xy=(0, 47), xycoords='axes points',
                     size=5, color='k', weight='bold', ha='right', va='top')

    figfile = filename.split('.')[0] + '.png'
    fig.savefig('./result/'+figfile, dpi=300)
    plt.close()


def get_efficiency(directory, file_name, spacer, base_pos_in_spacer, base_change):
    """
       Calculate editing efficiency of the edited base using the ratio
       of the edited base value to total value (edited + nonedited values)

       Input: directory of files, name of file, spacer sequence, position of
       base in spacer, base conversion, control or sample file, from csv or just one file

       Output: efficiency, result figure, result excel file
    """
    # delete the var directory
    sample = SeqIO.read('./' + directory + '/' + file_name, 'abi')
    spacer = spacer.upper()

    # Channels for bases: G, A, T, C
    trace_G = sample.annotations['abif_raw']['DATA9']
    trace_A = sample.annotations['abif_raw']['DATA10']
    trace_T = sample.annotations['abif_raw']['DATA11']
    trace_C = sample.annotations['abif_raw']['DATA12']
    trace = {
        'G': trace_G,
        'A': trace_A,
        'T': trace_T,
        'C': trace_C
    }
    
    # Base calls for file
    call = sample.annotations['abif_raw']['PBAS1']
    if isinstance(call,bytes):
        call = call.decode('utf-8')
    # These are the positions that correspond to each base call in the original trace
    peaks_vals = sample.annotations['abif_raw']['PLOC1']

    # Find the spacer position in the base call
    pos_in_call,strand,base_pos_in_spacer = find_subseq(spacer,call,base_change,base_pos_in_spacer)
    if strand == '-':
        spacer = rev_comp(spacer)
    if pos_in_call == -1:
        # Replace with raise or error
        print("*********Warning*********")
        print("Cannot find spacer " + spacer + " in " + file_name + ".")
        return None


    # Estimate the background for each base from the 101th base to the 50th of the last.
    bg_dict = background_estimation(call,peaks_vals,trace)

    # Determine the peak values of each base at each position along the spacer
    GG = []  # Store the peak intensity value of G at each position along the spacer
    AA = []
    TT = []
    CC = []
    G_vals = []  # Store the percent value of G at each position along the spacer
    A_vals = []
    T_vals = []
    C_vals = []

    #Calculate the peak values
    for i in range(0, len(spacer)):
        base_pos = peaks_vals[pos_in_call + i]

        # Obtain the peak value and substrat background
        # If the peak value is less than background, then make it 0.

        gg = max(max(trace_G[base_pos - 1: base_pos + 1]) - bg_dict['G'], 0)
        aa = max(max(trace_A[base_pos - 1: base_pos + 1]) - bg_dict['A'], 0)
        tt = max(max(trace_T[base_pos - 1: base_pos + 1]) - bg_dict['T'], 0)
        cc = max(max(trace_C[base_pos - 1: base_pos + 1]) - bg_dict['C'], 0)

        all_sum = gg + aa + tt + cc

        GG.append(int(gg))
        AA.append(int(aa))
        TT.append(int(tt))
        CC.append(int(cc))

        G_vals.append(gg * 100 / all_sum)
        A_vals.append(aa * 100 / all_sum)
        T_vals.append(tt * 100 / all_sum)
        C_vals.append(cc * 100 / all_sum)

        peak_intensity = {
            'A':AA,'G':GG,'C':CC,'T':TT
        }

    ratio_dict = {'G': G_vals[base_pos_in_spacer - 1], 'A': A_vals[base_pos_in_spacer - 1],
                  'T': T_vals[base_pos_in_spacer - 1], 'C': C_vals[base_pos_in_spacer - 1]}
    ratio = ratio_dict[base_change[1]]
    efficiency = "%.2f" % ratio

    data = [(list(spacer)), G_vals, A_vals,
           T_vals, C_vals]

    '''Output excel and figure'''
    write_excel(data,spacer,file_name,pos_in_call,base_pos_in_spacer,bg_dict,strand,peak_intensity)

    # Replace any 100% values to 99% for neat plotting of the table.
    lookup_map = {100: 99}
    pos_list = [map(round, vals) for vals in [G_vals, A_vals, T_vals, C_vals]]
    pos_list = [tuple(lookup_map.get(int(e), e) for e in s) for s in pos_list]
    plt_figure(trace,peaks_vals,spacer,pos_in_call,file_name,pos_list,base_pos_in_spacer,strand)

    return efficiency


def main():
    # Handle csv input with multiple samples
    # Example: python Beat8.py ./example/template.csv
    if 'result' not in os.listdir('./'):
        os.mkdir('result')
    if (sys.argv[1].endswith(".csv")):
        df = pd.read_csv(sys.argv[1])
        efficiencies = []

        for index, row in df.iterrows():
            directory = row['Directory']
            sample_file = row['Sample']
            target = row['Target Sequence']
            be_position = row['Base Position in Spacer']
            base_change = row['Conversion']

            print('processing: ' + sample_file)
            efficiency = get_efficiency(directory, sample_file, target, int(be_position), base_change)
            # samp_ratio = "%.3f" % ratio
            efficiencies.append(efficiency)

        # Create new column of efficiencies
        df['Efficiency'] = efficiencies

        df.to_csv(sys.argv[1], index=False)

    # One sample or entire folder
    # Example: python Beat8.py foldername example.ab1 GGgtaaatgtagttgataat 6 AG
    # Example: python Beat8.py foldername all GGgtaaatgtagttgataat 6 AG
    else:
        directory = sys.argv[1]
        sample_file = sys.argv[2]
        target = sys.argv[3]
        be_position = sys.argv[4]
        base_change = sys.argv[5]
        print('The directory is:', './' + directory)

        if sample_file == 'all' or sample_file == '':
            for files in os.listdir('./' + directory + '/'):
                if files[-4:] == '.ab1' or files.split('.')[1] == 'ab1':
                    sample_file = files
                    print('processing: ' + sample_file)
                    efficiency = get_efficiency(directory, sample_file, target, int(be_position), base_change)
        else:
            efficiency = get_efficiency(directory, sample_file, target, int(be_position), base_change)


if __name__ == "__main__":
    main()

stop = timeit.default_timer()
print('Completed in %.3f s' % (stop - start))
