from __future__ import division
import sys
import os
os.environ[ 'MPLCONFIGDIR' ] = '/tmp/'
#sys.path.append("/home/renzhi/.local/lib/python3.8/site-packages/");
from Bio import SeqIO
from Bio.Seq import Seq
import matplotlib
matplotlib.use('Agg')
#matplotlib.use('PS')
import matplotlib.pyplot as plt
from matplotlib.table import table
import matplotlib.colors as mcolors
import pandas as pd
import timeit
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from matplotlib.font_manager import FontProperties
import numpy as np
from scipy import stats

start = timeit.default_timer()
# spacer_err = []    # track the file which contain no spacer
#print("running...<br>")

#############################################################
# Beat - Base Editing Analysis Tool
# Beat determines editing effiency after substracting
#   the background noise without normalization to controls. 
# It finds the noise and filters the outliers from the noise
#   using the Median Absolute Deviation (MAD) method
#
# Run batch analysis as 
#   python Beat.py foldername all GAGTATGAGGCATAGACTGC 5 AG
#   to analyze all sequencing files if they have the same spacer
#
#   or python Beat.py ./data/template1.csv if they contain
#   different spacers
#
# To analyze individual sequencing file,
#   python Beat.py foldername Site2_PCR_420.ab1 GAGTATGAGGCATAGACTGC 5 AG
#############################################################

# define degenerate nucleotides


A = {"A"}
C = {"C"}
G = {"G"}
T = {"T"}

letter2nucl = {
    "A": A,
    "C": C,
    "G": G,
    "T": T,
    "R": A | G,
    "Y": C | T,
    "S": G | C,
    "W": A | T,
    "K": G | T,
    "M": A | C,
    "B": C | G | T,
    "D": A | G | T,
    "H": A | C | T,
    "V": A | C | G,
    "N": {"A", "C", "G", "T"}
    }


def rev_comp(seq):
    return str(Seq(seq).reverse_complement())

def removeOutliersMAD(x):
    # Median absolute deviation (MAD), based on the median, is a robust non-parametric statistics.
    # https://en.wikipedia.org/wiki/Median_absolute_deviation.
    x_median = np.median(x)
    resultList = np.copy(x)
    mad = 1.4826 * np.median(np.abs(x - x_median))
    resultList[(np.abs(x - x_median) > 3 * mad)] = x_median
    return resultList.tolist()


def removeOutliersZ(x):
    # The Z-score is the signed number of standard deviations by which the value of an observation
    # or data point is above the mean value of what is being observed or measured.
    z = np.abs(stats.zscore(x))
    resultList = []

    # See if we can del the for loop here
    for y in np.where(z ** 2 < 3):
        for yy in y:
            resultList.append(x[yy])

    return resultList


def removeOutliers(x, outlierConstant):
    # The interquartile range (IQR), also called the midspread or middle 50%, or technically H-spread,
    # is a measure of statistical dispersion, being equal to the difference between 75th and 25th
    # percentiles, or between upper and lower quartiles, IQR = Q3 - Q1.
    # It is a measure of the dispersion similar to standard deviation or variance,
    # but is much more robust against outliers.
    a = np.array(x)
    upper_quartile = np.percentile(a, 75)
    lower_quartile = np.percentile(a, 25)
    IQR = (upper_quartile - lower_quartile) * outlierConstant
    resultList = []
    for y in a.tolist():
        if lower_quartile - IQR >= y <= upper_quartile + IQR:
            resultList.append(y)
    return resultList


def make_colormap(seq):
    """Return a LinearSegmentedColormap
    seq: a sequence of floats and RGB-tuples. The floats should be increasing
    and in the interval (0,1).
    """
    seq = [(None,) * 3, 0.0] + list(seq) + [1.0, (None,) * 3]
    cdict = {'red': [], 'green': [], 'blue': []}
    for i, item in enumerate(seq):
        if isinstance(item, float):
            r1, g1, b1 = seq[i - 1]
            r2, g2, b2 = seq[i + 1]
            cdict['red'].append([item, r1, r2])
            cdict['green'].append([item, g1, g2])
            cdict['blue'].append([item, b1, b2])
    return mcolors.LinearSegmentedColormap('CustomMap', cdict)


def find_alignment(pattern, seq, start, end):

    EDNAFullLetters = ['A', 'T', 'G', 'C', 'S', 'W', 'R', 'Y', 'K', 'M', 'B', 'V', 'H', 'D', 'N', 'U']
    EDNAFullScores = [
        [5,-4,-4,-4,-4, 1, 3,-4,-4, 1,-4,-1,-1,-1,-2,-4],
        [-4, 5,-4,-4,-4, 1,-4, 3, 1,-4,-1,-4,-1,-1,-2, 5],
        [-4,-4, 5,-4, 1,-4, 3,-4, 1,-4,-1,-1,-4,-1,-2,-4],
        [-4,-4,-4, 5, 1,-4,-4, 3,-4, 1,-1,-1,-1,-4,-2,-4],
        [-4,-4, 1, 1,-1,-4,-2,-2,-2,-2,-1,-1,-3,-3,-1,-4],
        [1, 1,-4,-4,-4,-1,-2,-2,-2,-2,-3,-3,-1,-1,-1, 1],
        [1,-4, 1,-4,-2,-2,-1,-4,-2,-2,-3,-1,-3,-1,-1,-4],
        [-4, 1,-4, 1,-2,-2,-4,-1,-2,-2,-1,-3,-1,-3,-1, 1],
        [-4, 1, 1,-4,-2,-2,-2,-2,-1,-4,-1,-3,-3,-1,-1, 1],
        [1,-4,-4, 1,-2,-2,-2,-2,-4,-1,-3,-1,-1,-3,-1,-4],
        [-4,-1,-1,-1,-1,-3,-3,-1,-1,-3,-1,-2,-2,-2,-1,-1],
        [-1,-4,-1,-1,-1,-3,-1,-3,-3,-1,-2,-1,-2,-2,-1,-4],
        [-1,-1,-4,-1,-3,-1,-3,-1,-3,-1,-2,-2,-1,-2,-1,-1],
        [-1,-1,-1,-4,-3,-1,-1,-3,-1,-3,-2,-2,-2,-1,-1,-1],
        [-2,-2,-2,-2,-1,-1,-1,-1,-1,-1,-1,-1,-1,-1,-1,-2],
        [-4, 5,-4,-4,-4, 1,-4, 1, 1,-4,-1,-4,-1,-1,-2, 5]]

    EDNAFull = dict()
    for i, a in enumerate(EDNAFullLetters):
        for j, b in enumerate(EDNAFullLetters):
            EDNAFull[(a, b)] = EDNAFullScores[i][j]

    strand = '-'
    from Bio import pairwise2
    from Bio.pairwise2 import format_alignment
    seq = seq[start:end]
    #alignments = pairwise2.align.lobalms(seq, pattern, 5, -3, -50, -50) #match 5 points, mismatch -3 points, gap -50 points, extending -50 points.
    alignments = pairwise2.align.localds(seq, pattern, EDNAFull, -50, -50)
    top_aln = alignments[0]
    aln_seq, aln_spacer, score, begin1, end1 = top_aln

    #print(alignments)
    #print (aln_seq+'\n'+aln_spacer+' '+str(score)+'\n')

    pattern = rev_comp(pattern)
    alignments2 = pairwise2.align.localds(seq, pattern, EDNAFull, -50, -50)
    #alignments2 = pairwise2.align.localms(seq, pattern, 5, -3, -50, -50)
    top_aln2 = alignments2[0]
    aln_seq2, aln_spacer2, score2, begin2, end2 = top_aln2

    #print (aln_seq2+'\n'+aln_spacer2+' '+str(score2)+'\n')

    #print('score threshold: ' + str(4*len(pattern)) + '\n')
    if score > score2 and score > 3*len(pattern):
        strand = '+'
        pos_in_call = begin1 + start
    elif score2 > score and score2 > 3*len(pattern):
        strand = '-'
        pos_in_call = begin2 + start
    else:
        pos_in_call = -1

    return pos_in_call, strand

def background_estimation(base_call, peaks_vals, trace, start, end):

    Gbg_vals = []
    Abg_vals = []
    Tbg_vals = []
    Cbg_vals = []
    for i in range(start, end):
        x = base_call[i]
        base_pos = peaks_vals[i]
        if x in letter2nucl["N"]:
            if x != "A":
                Abg_vals.append(max(trace['A'][base_pos - 3: base_pos + 3]))
            if x != "G":
                Gbg_vals.append(max(trace['G'][base_pos - 3: base_pos + 3]))
            if x != "C":
                Cbg_vals.append(max(trace['C'][base_pos - 3: base_pos + 3]))
            if x != "T":
                Tbg_vals.append(max(trace['T'][base_pos - 3: base_pos + 3]))

    Gbg_vals_cleaned = removeOutliersMAD(Gbg_vals)
    Abg_vals_cleaned = removeOutliersMAD(Abg_vals)
    Tbg_vals_cleaned = removeOutliersMAD(Tbg_vals)
    Cbg_vals_cleaned = removeOutliersMAD(Cbg_vals)
    #print([Gbg_vals,Abg_vals,Tbg_vals,Cbg_vals])
    #print([Gbg_vals_cleaned,Abg_vals_cleaned,Tbg_vals_cleaned,Cbg_vals_cleaned])
    bg_dict = {
        'G': round(sum(Gbg_vals_cleaned) / len(Gbg_vals_cleaned),1),
        'A': round(sum(Abg_vals_cleaned) / len(Abg_vals_cleaned),1),
        'T': round(sum(Tbg_vals_cleaned) / len(Tbg_vals_cleaned),1),
        'C': round(sum(Cbg_vals_cleaned) / len(Cbg_vals_cleaned),1)
    }
    return bg_dict


def write_excel(directory, data, spacer, filename, spacer_pos, mutated_base, estimation, peak_intensity):

    # Excel output Construction
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Convert to 1 digit
    for i in data[1:]:
        for j in i:
            round(j,1)
    # Rows can also be appended
    for row in data:
        ws.append(row)
    maxRow = ws.max_row
    maxCol = ws.max_column

    # Format header row
    for cell in ws["1:1"]:
        cell.font = Font(size=12, color='00000000', italic=False, bold=True)
        cell.fill = PatternFill(fgColor='D3D3D3', fill_type='solid')

    # Format data row
    for x in range(2, maxRow + 1):
        for y in range(1, maxCol):
            ws.cell(row=x, column=y).font = Font(size=12, bold=False)

    # Highlight the mutated base
    ws.cell(row=1, column=mutated_base).font = Font(size=12, color='FF0000', bold=True)

    # Highlight the entire column of the mutated base
    for x in range(1, maxRow + 1):
        ws.cell(row=x, column=mutated_base).fill = PatternFill(fgColor='FFEE08', fill_type='solid')

    # Insert a column in the beginning
    ws.insert_cols(1)
    '''simplify this part'''
    # Add base index for each row
    ws.cell(row=1, column=1).value = 'Spacer'
    ws.cell(row=1, column=1).font = Font(size=12, color='000000', bold=True)
    ws.cell(row=2, column=1).value = 'G'
    ws.cell(row=2, column=1).font = Font(size=12, color='000000', bold=True)
    ws.cell(row=3, column=1).value = 'A'
    ws.cell(row=3, column=1).font = Font(size=12, color='008000', bold=True)
    ws.cell(row=4, column=1).value = 'T'
    ws.cell(row=4, column=1).font = Font(size=12, color='FF0000', bold=True)
    ws.cell(row=5, column=1).value = 'C'
    ws.cell(row=5, column=1).font = Font(size=12, color='0000FF', bold=True)
    for x in range(1, maxRow + 1):
        ws.cell(row=x, column=1).fill = PatternFill(fgColor='D3D3D3', fill_type='solid')

    # Insert information after the table
    ws.cell(row=7, column=1).value = filename
    ws.cell(row=8, column=1).value = 'spacer: ' + spacer
    ws.cell(row=9, column=1).value = 'spacer position in the base calls: ' + str(spacer_pos + 1)
    ws.cell(row=10, column=1).value = 'Estimated background: G:'+str(estimation['G'])+'; A:'+str(estimation['A'])+\
                                      '; T:'+str(estimation['T'])+'; C:'+str(estimation['C'])
    ws.cell(row=11, column=1).value = 'Peak values at position ' + str(mutated_base) + ', G:' + str(
        peak_intensity['G'][mutated_base - 1]) + '; A:' + str(peak_intensity['A'][mutated_base - 1]) + '; T: ' + str(
        peak_intensity['T'][mutated_base - 1]) + '; C: ' + str(peak_intensity['C'][mutated_base - 1])

    # Save the file
    wb.save(directory+'/result/'+filename.split('.ab1')[0]+'.xlsx')


def plt_figure(directory, trace, peaks_vals, spacer, pos_in_call, filename, pos_list, base_pos_in_spacer):
    # Plot the trace and table showing % of each base along the spacer

    fig, (ax1, ax2) = plt.subplots(2, figsize=(4, 2.5))
    fig.patch.set_visible(True)
    ax1 = plt.subplot2grid((2, 1), (0, 0))
    ax2 = plt.subplot2grid((2, 1), (1, 0))
    Arialfont = {'fontname': 'Arial'}
    ax1.plot(trace['G'][peaks_vals[pos_in_call] - 8: peaks_vals[pos_in_call + len(spacer) - 1] + 8], color='black',
             label='G', linewidth=0.6)
    ax1.plot(trace['A'][peaks_vals[pos_in_call] - 8: peaks_vals[pos_in_call + len(spacer) - 1] + 8], color='g', label='A',
             linewidth=0.6)  # A
    ax1.plot(trace['T'][peaks_vals[pos_in_call] - 8: peaks_vals[pos_in_call + len(spacer) - 1] + 8], color='r', label='T',
             linewidth=0.6)  # T
    ax1.plot(trace['C'][peaks_vals[pos_in_call] - 8: peaks_vals[pos_in_call + len(spacer) - 1] + 8], color='b', label='C',
             linewidth=0.6)  # C
    ax1.annotate('G', xy=(10, 60), xycoords='axes points', size=5, color='black', weight='bold', ha='right', va='top')
    ax1.annotate('A', xy=(14, 60), xycoords='axes points', size=5, color='g', weight='bold', ha='right', va='top')
    ax1.annotate('T', xy=(18, 60), xycoords='axes points',
                 size=5, color='r', weight='bold', ha='right', va='top')
    ax1.annotate('C', xy=(22, 60), xycoords='axes points',
                 size=5, color='b', weight='bold', ha='right', va='top')
    ax1.axis('off')
    ax1.axis('tight')
    ax1.set_title(filename.split(".ab1")[0], fontsize=8, **Arialfont)

    # Put a legend to the right of the current axis
    # plt.legend([a_plot, c_plot, g_plot, t_plot], ['A', 'C', 'G', 'T'], bbox_to_anchor=(0.98,0.5), loc="center left", prop={'size': 6}, borderaxespad=0., ncol=1)

    ax2.axis('off')
    ax2.axis('tight')
    df = pd.DataFrame(pos_list, columns=list(spacer))
    xx = (df.values * 2.56).astype(int)
    c = mcolors.ColorConverter().to_rgb
    mycm = make_colormap(
        [c('white'), c('yellow'), 0.3, c('yellow'), c('red'), 0.60, c('red'), c('palegreen'), 0.8, c('palegreen')])

    # colours=plt.cm.CMRmap(xx)
    colours = plt.cm.get_cmap(mycm)(xx)
    stats_table = table(ax2, cellText=df.values, rowLabels='GATC', colLabels=df.columns, loc='center', cellLoc='center',
                        cellColours=colours)
    stats_table._cells[(0, base_pos_in_spacer - 1)]._text.set_color('red')
    stats_table._cells[(2, -1)]._text.set_color('green')
    stats_table._cells[(3, -1)]._text.set_color('red')
    stats_table._cells[(4, -1)]._text.set_color('blue')
    # stats_table._cells[(1,base_pos_in_spacer-1)].set_facecolor('red')

    for (row, col), cell in stats_table.get_celld().items():
        cell.set_linewidth(0)
        if (col == -1):
            cell.set_text_props(fontproperties=FontProperties(family='Arial', size=8, weight='bold'))
        elif (row == 0):
            cell.set_text_props(fontproperties=FontProperties(family='Arial', size=8))
        else:
            cell.set_text_props(fontproperties=FontProperties(family='Arial', size=6))

            # shrink current axis
    box = ax2.get_position()
    ax2.set_position([box.x0 + 0.038, box.y0 + 0.15, box.width * 0.9, box.height * 0.8])
    # ax2.arrow(0, 20, 10, 47, head_width=1, head_length=2, fc='k', ec='k')

    ax2.annotate("5'>", xy=(0, 47), xycoords='axes points',
                     size=5, color='k', weight='bold', ha='right', va='top')


    figfile = filename.split('.ab1')[0] + '.png'
    fig.savefig(directory+'/result/'+figfile, dpi=300)
    plt.close()


def get_efficiency(directory, file_name, spacer, base_pos_in_spacer, start, end):
    """
       Calculate editing efficiency of the edited base using the ratio
       of the edited base value to total value (edited + nonedited values)
       Input: directory of files, name of file, spacer sequence, position of
       base in spacer, base conversion, control or sample file, from csv or just one file
       Output: efficiency, result figure, result excel file
    """
    sample = SeqIO.read(directory + '/' + file_name, 'abi')

    spacer = spacer.upper()

    # Channels for bases: G, A, T, C
    trace_G = sample.annotations['abif_raw']['DATA9']
    trace_A = sample.annotations['abif_raw']['DATA10']
    trace_T = sample.annotations['abif_raw']['DATA11']
    trace_C = sample.annotations['abif_raw']['DATA12']
    trace = {
        'G': trace_G,
        'A': trace_A,
        'T': trace_T,
        'C': trace_C
    }

    # Base calls for file
    call = sample.annotations['abif_raw']['PBAS1']
    if isinstance(call,bytes):
        call = call.decode('utf-8')
    # Per-base quality
    # quality = [ord(ch) for ch in str(sample.annotations['abif_raw']['PCON1']) ]
    # Peak area ratio
    # peakratio = sample.annotations['abif_raw']['phAR1']

    # print(call)
    # print(quality)
    # print(peakratio)

    # These are the positions that correspond to each base call in the original trace
    peaks_vals = sample.annotations['abif_raw']['PLOC1']

    # Find the spacer position in the base call
    # pos_in_call,strand = find_subseq(spacer,call,base_change,start,end)
    pos_in_call,strand = find_alignment(spacer,call,start,end) #find_alignment(pattern, seq, start, end)

    if strand == "-":
        rc_trace_G = trace_C[::-1]
        rc_trace_C = trace_G[::-1]
        rc_trace_A = trace_T[::-1]
        rc_trace_T = trace_A[::-1]
        #print('trace_G length: ', len(trace_G))
        new_peaks_vals = tuple([len(trace_G) - i for i in peaks_vals])
        rc_peaks_vals = new_peaks_vals[::-1]
        rc_call = rev_comp(call)
        rc_start = len(call) - end + 1
        rc_end = len(call) - start + 1
        rc_trace = {
            'G': rc_trace_G,
            'A': rc_trace_A,
            'T': rc_trace_T,
            'C': rc_trace_C }

    if pos_in_call == -1:
        # Replace with raise or error
        print("*********Warning*********")
        print("Cannot find spacer " + spacer + " in " + file_name + ".")
        return None

    # Estimate the background for each base from the 101th base to the 50th of the last.
    bg_dict = background_estimation(call,peaks_vals,trace,start,end)
    #print(bg_dict['G'], bg_dict['A'], bg_dict['T'], bg_dict['C'])

    # Determine the peak values of each base at each position along the spacer
    GG = []  # Store the peak intensity value of G at each position along the spacer
    AA = []
    TT = []
    CC = []
    G_vals = []  # Store the percent value of G at each position along the spacer
    A_vals = []
    T_vals = []
    C_vals = []

    #Calculate the peak values
    for i in range(0, len(spacer)):
        base_pos = peaks_vals[pos_in_call + i]

        #print(pos_in_call, base_pos, trace_T[base_pos])

        # Obtain the peak value and substrat background
        # If the peak value is less than background, then make it 0.

        if strand == '+':
            gg = max(max(trace_G[base_pos - 1: base_pos + 1]) - bg_dict['G'], 0)
            aa = max(max(trace_A[base_pos - 1: base_pos + 1]) - bg_dict['A'], 0)
            tt = max(max(trace_T[base_pos - 1: base_pos + 1]) - bg_dict['T'], 0)
            cc = max(max(trace_C[base_pos - 1: base_pos + 1]) - bg_dict['C'], 0)

        if strand == '-':
            gg = max(max(trace_C[base_pos - 1: base_pos + 1]) - bg_dict['C'], 0)
            aa = max(max(trace_T[base_pos - 1: base_pos + 1]) - bg_dict['T'], 0)
            tt = max(max(trace_A[base_pos - 1: base_pos + 1]) - bg_dict['A'], 0)
            cc = max(max(trace_G[base_pos - 1: base_pos + 1]) - bg_dict['G'], 0)

        #print(gg, aa, tt, cc)

        all_sum = gg + aa + tt + cc

        GG.append(int(gg))
        AA.append(int(aa))
        TT.append(int(tt))
        CC.append(int(cc))

        G_vals.append(gg * 100 / all_sum)
        A_vals.append(aa * 100 / all_sum)
        T_vals.append(tt * 100 / all_sum)
        C_vals.append(cc * 100 / all_sum)


    if strand == '-':
        GG = GG[::-1]
        AA = AA[::-1]
        TT = TT[::-1]
        CC = CC[::-1]
        G_vals = G_vals[::-1]
        A_vals = A_vals[::-1]
        T_vals = T_vals[::-1]
        C_vals = C_vals[::-1]
        trace = rc_trace
        peaks_vals = rc_peaks_vals
        pos_in_call = len(call) - pos_in_call - len(spacer)


    peak_intensity = {'A':AA,'G':GG,'C':CC,'T':TT}
    ratio_dict = {'G': G_vals[base_pos_in_spacer - 1], 'A': A_vals[base_pos_in_spacer - 1],
                  'T': T_vals[base_pos_in_spacer - 1], 'C': C_vals[base_pos_in_spacer - 1]}
    #ratio = ratio_dict[base_change[1]]
    #efficiency = "%.2f" % ratio

    data = [(list(spacer)), G_vals, A_vals, T_vals, C_vals]
    '''Output excel and figure'''
    write_excel(directory,data,spacer,file_name,pos_in_call,base_pos_in_spacer,bg_dict,peak_intensity)

    # Replace any 100% values to 99% for neat plotting of the table.
    lookup_map = {100: 99}
    pos_list = [map(round, vals) for vals in [G_vals, A_vals, T_vals, C_vals]]
    pos_list = [tuple(lookup_map.get(int(e), e) for e in s) for s in pos_list]
    plt_figure(directory,trace,peaks_vals,spacer,pos_in_call,file_name,pos_list,base_pos_in_spacer)

    df = pd.DataFrame([G_vals, A_vals, T_vals, C_vals], columns=list(spacer), index=['G', 'A', 'T', 'C'])
    #df.to_csv(directory+'/result/'+file_name.split('.ab1')[0]+'.csv')
    #return efficiency


def main():
    # Handle csv input with multiple samples
    # Example: python Beat.py ./example/template.csv
    if 'result' not in os.listdir(sys.argv[1]):
        os.mkdir(sys.argv[1]+'/result')

    directory = sys.argv[1]
    sample_file = sys.argv[2]
    target = sys.argv[3]
    be_position = sys.argv[4]
    #base_change = sys.argv[5]
    start_pos = sys.argv[5]
    end_pos = sys.argv[6]
    #print('The directory is:', directory, '<br>')

    if sample_file == 'all' or sample_file == '':
            for files in os.listdir(directory):
            # for files in os.listdir('./' + directory + '/'):
                if files[-4:] == '.ab1':
                    sample_file = files
                    #print('processing: ' + sample_file + '<br>')
                    efficiency = get_efficiency(directory, sample_file, target, int(be_position), int(start_pos), int(end_pos))
    else:
            efficiency = get_efficiency(directory, sample_file, target, int(be_position), int(start_pos), int(end_pos))


if __name__ == "__main__":
    main()

stop = timeit.default_timer()
print('Completed in %.3f s' % (stop - start), ' <br>')
